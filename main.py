import win32com.client
import re
from datetime import datetime, timedelta, date
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import defaultdict
import win32timezone

def get_emails(from_date, subject_keyword):
    try:
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        inbox = outlook.GetDefaultFolder(6)  # 6 refers to the Inbox

        # **New Code Starts Here**
        # Access the "Schedule Changes" subfolder within the Inbox
        schedule_changes_folder = None
        for folder in inbox.Folders:
            if folder.Name.lower() == "schedule changes":
                schedule_changes_folder = folder
                break

        if not schedule_changes_folder:
            print('Subfolder "Schedule Changes" not found in Inbox.')
            return []

        messages = schedule_changes_folder.Items
        # **New Code Ends Here**

        messages.Sort("[SentOn]", Descending=False)

        restrict_date = from_date.strftime('%m/%d/%Y 12:00 AM')
        restricted = messages.Restrict(f"[SentOn] >= '{restrict_date}'")

        filtered_emails = []
        for email in restricted:
            try:
                if subject_keyword.lower() in email.Subject.lower():
                    filtered_emails.append(email)
            except AttributeError:
                continue

        return filtered_emails

    except Exception as e:
        print(f"Error fetching emails: {e}")
        return []

def parse_email(email_body):
    sent_date_match = re.search(r"Sent:\s+(.*)", email_body, re.IGNORECASE)
    if sent_date_match:
        sent_date_str = sent_date_match.group(1).strip()
        sent_datetime = None
        date_formats = [
            '%A, %B %d, %Y %I:%M %p',
            '%B %d, %Y %I:%M %p',
            '%a, %b %d, %Y %I:%M %p',
            '%d %B %Y %I:%M %p',
        ]
        for fmt in date_formats:
            try:
                sent_datetime = datetime.strptime(sent_date_str, fmt)
                break
            except ValueError:
                continue

        sent_date = sent_datetime.date() if sent_datetime else None
    else:
        sent_date = None

    additions_match = re.search(r"Additions:\s*(.*?)\s*(Cancellations:|Instructor Changes:|$)", email_body, re.IGNORECASE | re.DOTALL)
    cancellations_match = re.search(r"Cancellations:\s*(.*?)\s*(Instructor Changes:|$)", email_body, re.IGNORECASE | re.DOTALL)
    instructor_changes_match = re.search(r"Instructor Changes:\s*(.*)", email_body, re.IGNORECASE | re.DOTALL)

    def count_valid_lines(section_text, pattern=None):
        if not section_text:
            return 0
        lines = section_text.strip().split("\n")
        if pattern:
            return len([line for line in lines if line.strip() and re.match(pattern, line.strip())])
        else:
            return len([line for line in lines if line.strip()])

    additions_pattern = r".+\s*/\s*CRN\s*\d+"
    cancellations_pattern = r".+\s*/\s*CRN\s*\d+"
    instructor_pattern = r".+\s*/\s*CRN\s*\d+\s*-\s*.+"

    additions = count_valid_lines(additions_match.group(1) if additions_match else None, additions_pattern)
    cancellations = count_valid_lines(cancellations_match.group(1) if cancellations_match else None, cancellations_pattern)
    instructor_changes = count_valid_lines(
        instructor_changes_match.group(1) if instructor_changes_match else None,
        instructor_pattern
    )

    return sent_date, additions, cancellations, instructor_changes

def write_to_excel(data, weekly_data, total_counts, averages, email_details, output_file):
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws_daily = wb.active
    ws_daily.title = "Daily Changes Summary"

    headers = ["Date", "Section Additions", "Section Cancellations", "Instructor Changes", "Note"]
    ws_daily.append(headers)
    for col in range(1, 6):
        cell = ws_daily.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    sorted_dates = sorted(data.keys())

    for date_obj in sorted_dates:
        additions = data[date_obj]['additions']
        cancellations = data[date_obj]['cancellations']
        instructor_changes = data[date_obj]['instructor_changes']
        email_count = data[date_obj]['email_count']
        note = "Multiple emails processed" if email_count > 1 else ""
        ws_daily.append([date_obj, additions, cancellations, instructor_changes, note])

    for row_cells in ws_daily.iter_rows(min_row=2, max_row=ws_daily.max_row, min_col=1, max_col=5):
        for cell in row_cells:
            cell.alignment = center_alignment
            cell.border = thin_border
            if isinstance(cell.value, datetime) or isinstance(cell.value, date):
                cell.number_format = 'mmm dd yyyy'

    for column_cells in ws_daily.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        adjusted_width = (max_length + 4)
        column_letter = column_cells[0].column_letter
        ws_daily.column_dimensions[column_letter].width = adjusted_width

    ws_daily.freeze_panes = "A2"

    ws_weekly = wb.create_sheet(title="Weekly Changes Summary")

    headers = ["Week Start Date", "Week End Date", "Date Range", "Section Additions", "Section Cancellations", "Instructor Changes"]
    ws_weekly.append(headers)
    for col in range(1, 7):
        cell = ws_weekly.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    sorted_weeks = sorted(weekly_data.keys())

    for week_start_date in sorted_weeks:
        week_end_date = week_start_date + timedelta(days=6)
        date_range = f"{week_start_date.strftime('%b %d, %Y')} - {week_end_date.strftime('%b %d, %Y')}"
        additions = weekly_data[week_start_date]['additions']
        cancellations = weekly_data[week_start_date]['cancellations']
        instructor_changes = weekly_data[week_start_date]['instructor_changes']
        ws_weekly.append([week_start_date, week_end_date, date_range, additions, cancellations, instructor_changes])

    for row_cells in ws_weekly.iter_rows(min_row=2, max_row=ws_weekly.max_row, min_col=1, max_col=6):
        for cell in row_cells:
            cell.alignment = center_alignment
            cell.border = thin_border
            if cell.column in [1, 2]:
                cell.number_format = 'mmm dd yyyy'

    for column_cells in ws_weekly.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        adjusted_width = (max_length + 4)
        column_letter = column_cells[0].column_letter
        ws_weekly.column_dimensions[column_letter].width = adjusted_width

    ws_weekly.freeze_panes = "A2"

    ws_stats = wb.create_sheet(title="Overall Statistics")

    ws_stats.cell(row=1, column=1, value="Total Counts")
    ws_stats.cell(row=1, column=1).font = header_font
    ws_stats.cell(row=1, column=1).fill = header_fill
    ws_stats.cell(row=1, column=1).alignment = center_alignment
    ws_stats.cell(row=1, column=1).border = thin_border
    ws_stats.cell(row=1, column=2).border = thin_border

    row = 2
    for key, value in total_counts.items():
        ws_stats.cell(row=row, column=1, value=key)
        ws_stats.cell(row=row, column=1).font = Font(bold=True)
        ws_stats.cell(row=row, column=1).alignment = Alignment(horizontal="left")
        ws_stats.cell(row=row, column=1).border = thin_border

        ws_stats.cell(row=row, column=2, value=value)
        ws_stats.cell(row=row, column=2).alignment = center_alignment
        ws_stats.cell(row=row, column=2).border = thin_border
        row += 1

    ws_stats.cell(row=row, column=1, value="Averages per Day")
    ws_stats.cell(row=row, column=1).font = header_font
    ws_stats.cell(row=row, column=1).fill = header_fill
    ws_stats.cell(row=row, column=1).alignment = center_alignment
    ws_stats.cell(row=row, column=1).border = thin_border
    ws_stats.cell(row=row, column=2).border = thin_border
    row += 1

    for key, value in averages.items():
        ws_stats.cell(row=row, column=1, value=key)
        ws_stats.cell(row=row, column=1).font = Font(bold=True)
        ws_stats.cell(row=row, column=1).alignment = Alignment(horizontal="left")
        ws_stats.cell(row=row, column=1).border = thin_border

        ws_stats.cell(row=row, column=2, value=round(value, 2))
        ws_stats.cell(row=row, column=2).alignment = center_alignment
        ws_stats.cell(row=row, column=2).border = thin_border
        row += 1

    for column_cells in ws_stats.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        adjusted_width = (max_length + 4)
        column_letter = column_cells[0].column_letter
        ws_stats.column_dimensions[column_letter].width = adjusted_width

    ws_stats.freeze_panes = "A2"

    ws_emails = wb.create_sheet(title="Email Details")

    headers = ["Email Sent On", "Subject", "Parsed Sent Date", "Section Additions", "Section Cancellations", "Instructor Changes"]
    ws_emails.append(headers)
    for col in range(1, 7):
        cell = ws_emails.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    for email in email_details:
        ws_emails.append([
            email['email_sent_on'],
            email['subject'],
            email['sent_date'],
            email['additions'],
            email['cancellations'],
            email['instructor_changes']
        ])

    for row_cells in ws_emails.iter_rows(min_row=2, max_row=ws_emails.max_row, min_col=1, max_col=6):
        for cell in row_cells:
            cell.alignment = center_alignment
            cell.border = thin_border
            if isinstance(cell.value, datetime):
                cell.number_format = 'mmm dd yyyy hh:mm AM/PM'
            elif isinstance(cell.value, date):
                cell.number_format = 'mmm dd yyyy'

    for column_cells in ws_emails.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        adjusted_width = (max_length + 4)
        column_letter = column_cells[0].column_letter
        ws_emails.column_dimensions[column_letter].width = adjusted_width

    ws_emails.freeze_panes = "A2"

    try:
        wb.save(output_file)
        wb.close()
        print(f"Data successfully written to {output_file}")
    except Exception as e:
        print(f"Failed to save the workbook: {e}")

if __name__ == "__main__":
    from_date = datetime(2024, 11, 15)
    subject_keyword = "Changes to the 202510 Course Schedule"
    output_file = "Course_Changes_Summary.xlsx"

    emails = get_emails(from_date, subject_keyword)
    if not emails:
        print("No matching emails found.")
    else:
        data = defaultdict(lambda: {'additions': 0, 'cancellations': 0, 'instructor_changes': 0, 'email_count': 0})
        email_details = []

        for email in emails:
            try:
                sent_date, additions, cancellations, instructor_changes = parse_email(email.Body)

                if sent_date is None:
                    sent_date = email.SentOn.date()

                data[sent_date]['additions'] += additions
                data[sent_date]['cancellations'] += cancellations
                data[sent_date]['instructor_changes'] += instructor_changes
                data[sent_date]['email_count'] += 1

                email_sent_on_naive = email.SentOn.replace(tzinfo=None) if email.SentOn.tzinfo else email.SentOn

                email_details.append({
                    'sent_date': sent_date,
                    'email_sent_on': email_sent_on_naive,
                    'subject': email.Subject,
                    'additions': additions,
                    'cancellations': cancellations,
                    'instructor_changes': instructor_changes,
                })

            except Exception as e:
                print(f"Error processing an email: {e}")
                continue

        if data:
            data = dict(data)

            total_additions = sum(entry['additions'] for entry in data.values())
            total_cancellations = sum(entry['cancellations'] for entry in data.values())
            total_instructor_changes = sum(entry['instructor_changes'] for entry in data.values())

            days_count = len(data)

            avg_additions = total_additions / days_count if days_count else 0
            avg_cancellations = total_cancellations / days_count if days_count else 0
            avg_instructor_changes = total_instructor_changes / days_count if days_count else 0

            total_counts = {
                'Total Additions': total_additions,
                'Total Cancellations': total_cancellations,
                'Total Instructor Changes': total_instructor_changes
            }

            averages = {
                'Average Additions per Day': avg_additions,
                'Average Cancellations per Day': avg_cancellations,
                'Average Instructor Changes per Day': avg_instructor_changes
            }

            weekly_data = defaultdict(lambda: {'additions': 0, 'cancellations': 0, 'instructor_changes': 0})

            for date_obj, counts in data.items():
                week_start_date = date_obj - timedelta(days=date_obj.weekday())
                weekly_data[week_start_date]['additions'] += counts['additions']
                weekly_data[week_start_date]['cancellations'] += counts['cancellations']
                weekly_data[week_start_date]['instructor_changes'] += counts['instructor_changes']

            weekly_data = dict(weekly_data)

            write_to_excel(data, weekly_data, total_counts, averages, email_details, output_file)
        else:
            print("No data to write after processing emails.")
